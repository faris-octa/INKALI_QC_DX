from os.path import abspath

import streamlit as st
from openpyxl import load_workbook

# open workbook
EXCEL_FILE = abspath('C:/Users/INKALI-PC/project/INKALI_QC_DX/PEG 600DO-IK/av_data.xlsx')
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# initiate table
header = [cell.value for cell in ws[1]]
data = []

st.set_page_config(page_title="Faris' Webpage", page_icon=":tada:", layout="wide")

st.table(header)

