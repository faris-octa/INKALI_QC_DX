from os.path import abspath
import PySimpleGUI as sg
from openpyxl import load_workbook
from datetime import datetime

# AV Standard
std = 7.0

# initiate excel file
EXCEL_FILE = abspath('C:/Users/INKALI-PC/project/INKALI_QC_DX/PEG 600DO-IK/pegav.xlsx')
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# initiate data structure
header = [cell.value for cell in ws[1]]
data = []

#### Layout ####
header_row = [
    [sg.Text('Acid Value Calculator', font=('Arial', 20), justification='center')],
    [sg.Text('INKALI QC', font=('Arial', 15), justification='center')]
]

table_column = [
    # LOT, Step, Waktu, Operator QC, Reaksi (°C), Berat Sample (gr), Jumlah Titran (mL) ,Faktor Buret,
    # Faktor NaOH, AV, Instruksi
    [sg.Table(values=data, headings=header, 
              col_widths=[5,5,10,10,10,14,14,10,10,10,25], 
              auto_size_columns = False,
              visible_column_map=[False, True, True, False, True, True, True, False, False, True, False],
              justification = 'center',
              background_color = 'black',
              key='-TABLE-')],
    [sg.Submit(pad=(300, 5))]
]

input_column = [
    [sg.Frame('Input Sample',[
    [sg.Text('Lot:', size=14), sg.InputText(key='-LOT-', size=(12,1))],
    [sg.Text('Operator QC:', size=14), sg.InputText(key='-OPERATOR-', size=(12,1))],
    [sg.Text('Faktor Buret:', size=14), sg.InputText(key='-FAKTOR-BURET-', size=(12,1))],
    [sg.Text('Faktor NaOH:', size=14), sg.InputText(key='-FAKTOR-NaOH-', size=(12,1),)],
    ])],
    [sg.Text('Reaksi (\xb0C):', size=15), sg.Combo(values=('', 'Packing'), default_value='', readonly=False, size=(10,1), k='-REAKSI-')],
    [sg.Text('Berat Sample (gr):', size=15), sg.InputText(key='-BERAT-SAMPLE-', size=(12,1))],
    [sg.Text('Jumlah Titran (mL):', size=15), sg.InputText(key='-JUMLAH-TITRAN-', size=(12,1))],
    [sg.Button('Clear'), sg.Button('Add')]
]

layout = [
    [sg.Column(header_row)],
    [sg.HSeparator()],
    [
        sg.Column(input_column),
        sg.VSeperator(),
        sg.Column(table_column, vertical_alignment='Top'),
    ]
]

#### functions ####
def add_row():
    new_row = [lot, step, waktu, operator, reaksi, berat_sample, 
                jumlah_titran, faktor_buret, faktor_NaOH, AV, instruksi]
    data.append(new_row)
    window['-REAKSI-']('')
    window['-BERAT-SAMPLE-']('')
    window['-JUMLAH-TITRAN-']('')

#### Window ####
window = sg.Window('Acid Value Form (S1200-IK)', layout, keep_on_top=False, finalize=True, resizable=True)
window.set_min_size(window.size)

while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED, None):
        break
    
    if event == 'Submit' and len(data) != 0:
        for row in data:
            ws.append(row)
        wb.save(EXCEL_FILE)
        # print(data) <- uncomment if want to debugging
        sg.popup('Data saved!')
        for key in values:
            window[key]('')
        window.close()
  
    if event == 'Clear':
        for key in values:
            if key != '-TABLE-':
                window[key]('')
    
    if event == 'Add':
        try:
            lot = values['-LOT-']
            step = len(data) + 1
            waktu = str(datetime.now().strftime("%m/%d/%Y %H:%M"))
            operator = values['-OPERATOR-']
            reaksi = values['-REAKSI-']
            berat_sample = round(float(values['-BERAT-SAMPLE-']),2)
            jumlah_titran = float(values['-JUMLAH-TITRAN-'])
            faktor_buret = float(values['-FAKTOR-BURET-'])
            faktor_NaOH = float(values['-FAKTOR-NaOH-'])
            AV = round((jumlah_titran * faktor_buret * faktor_NaOH * 5.61) / berat_sample, 4)
            
            if reaksi.lower() == 'packing':
                if  0< AV < std:
                    instruksi = 'OK'
                else:
                    instruksi = 'NG'
                add_row()

            elif 30 < float(reaksi) < 100 or 170 <= float(reaksi) < 200:
                instruksi = 'lakukan pemanasan hingga 225\xb0C'
                add_row()

            elif 100 <= float(reaksi) < 170:
                if 1 < AV < std:
                    instruksi = 'Packing'
                else:
                    instruksi = 'Hubungi atasan'
                add_row()

            elif 200 <= float(reaksi) < 240:
                if AV > std:
                    instruksi = 'Tambah waktu pemanasan'
                else:
                    instruksi = 'Lakukan cooling hingga 120\xb0C'
                add_row()
            
            window['-TABLE-'].update(values=data)
        except ValueError as e:
            sg.PopupError(str('Mohon input data dengan benar'))

window.close()
